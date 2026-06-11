"""정밀 필터링된 요구사항 75건을 Excel 리포트로."""
import csv, json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

REPORT = Path(r'C:\Users\heen1\Desktop\assist\rfp_report')
SRC = REPORT / 'requirements_filtered.csv'
OUT = REPORT / 'RFP_요구사항_정밀필터.xlsx'

SMELL_COLS = ['모호어','수동태','복합의무','정량부재','미정의약어','주체모호','약한의무']

H_FILL = PatternFill('solid', fgColor='1F4E79')
H_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
TITLE_FONT = Font(bold=True, color='1F4E79', name='Arial', size=14)
SUB_FONT = Font(italic=True, color='606060', name='Arial', size=10)
CELL_FONT = Font(name='맑은 고딕', size=10)
SMELLY_FILL = PatternFill('solid', fgColor='FFF4CC')

wb = Workbook()

rows = list(csv.DictReader(open(SRC, encoding='utf-8-sig')))
summary = json.load(open(REPORT / 'requirements_summary.json', encoding='utf-8'))

# ───────── Sheet 1: 안내 ─────────
ws0 = wb.active
ws0.title = '안내'
ws0['A1'] = f"RFP 56건 — 정밀 필터링된 한국어 요구사항 ({summary['total_requirements']}건)"
ws0['A1'].font = TITLE_FONT
ws0.merge_cells('A1:C1')
ws0['A2'] = '※ 원래 추출 3,210 문장 중 요구사항 시그널만 통과시킨 결과'
ws0['A2'].font = SUB_FONT
ws0.merge_cells('A2:C2')

ws0['A4'] = '처리 단계'; ws0['B4'] = '결과'; ws0['C4'] = '설명'
for col in 'ABC':
    c = ws0[f'{col}4']; c.font = H_FONT; c.fill = H_FILL
    c.alignment = Alignment(horizontal='center')

steps = [
    ('1차 원문 추출 문장', summary['filtered_from'], '문장 분리 + 한국어 비율 ≥25%'),
    ('요구사항 1차 필터', '330', '의무 종결어("~여야 한다" 등) + 노이즈 차단'),
    ('사이트 공통 문장 제거', f"{330 - 78}", '5+ 사업에서 동일하게 등장 = 공통 푸터'),
    ('사업 내 중복 제거', summary['total_requirements'], '같은 사업 안의 중복 한 줄로'),
    ('Smell 검출', f"{summary['total_with_smell']} ({100*summary['total_with_smell']/summary['total_requirements']:.0f}%)", '진짜 요구사항 중 품질 문제 보유'),
    ('요구사항 있는 사업', f"{summary['projects_with_reqs']} / 56", '나머지는 본문 미확보 또는 표준 공고문구 위주'),
]
for i, (a, b, c) in enumerate(steps, 5):
    ws0[f'A{i}'] = a; ws0[f'B{i}'] = b; ws0[f'C{i}'] = c
    for col in 'ABC': ws0[f'{col}{i}'].font = CELL_FONT

ws0['A12'] = 'Smell 유형별'; ws0['A12'].font = H_FONT; ws0['A12'].fill = H_FILL
ws0['B12'] = '건수'; ws0['B12'].font = H_FONT; ws0['B12'].fill = H_FILL
SMELL_DESC = {
    '모호어':     '"필요한·적절한·신속히" 등 정량성 결여 표현',
    '수동태':     '"~되어야 한다" — 행위 주체 흐릿',
    '복합의무':   '한 문장에 2개 이상 의무 표현',
    '정량부재':   '성능 키워드인데 숫자 없음',
    '미정의약어': '정의 없이 등장하는 영문 약어',
    '주체모호':   '주어 없이 의무문 시작',
    '약한의무':   '"~한다/된다" 평서형 (해야 한다 권장)',
}
ws0['C12'] = '설명'; ws0['C12'].font = H_FONT; ws0['C12'].fill = H_FILL
for i, c in enumerate(SMELL_COLS, 13):
    ws0[f'A{i}'] = c; ws0[f'B{i}'] = summary[c]; ws0[f'C{i}'] = SMELL_DESC[c]
    for col in 'ABC': ws0[f'{col}{i}'].font = CELL_FONT

ws0.column_dimensions['A'].width = 28
ws0.column_dimensions['B'].width = 16
ws0.column_dimensions['C'].width = 70

# ───────── Sheet 2: 요구사항 ─────────
ws1 = wb.create_sheet('요구사항')
ws1['A1'] = f'정밀 필터링된 요구사항 ({len(rows)}건)'
ws1['A1'].font = TITLE_FONT
ws1.merge_cells('A1:L1')

HEADERS = ['No.', '기관', '사업명', '요구사항', 'Smell'] + SMELL_COLS + ['모호어 단어']
for c, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=3, column=c, value=h)
    cell.font = H_FONT; cell.fill = H_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

def parse_pid(pid):
    return pid.split('_', 1)[0] if '_' in pid else pid

for ri, r in enumerate(rows, 4):
    cells = [
        parse_pid(r['project_id']),
        r['org'], r['project'], r['sentence'],
        int(r['has_smell']),
    ] + [int(r[c]) for c in SMELL_COLS] + [r.get('vague_hits', '')]
    for c, v in enumerate(cells, 1):
        cell = ws1.cell(row=ri, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws1.row_dimensions[ri].height = 60

last = 3 + len(rows)
ws1.auto_filter.ref = f'A3:M{last}'
ws1.freeze_panes = 'A4'
# has_smell=1 행 노란 배경
ws1.conditional_formatting.add(
    f'A4:M{last}',
    FormulaRule(formula=[f'$E4=1'], fill=SMELLY_FILL)
)
widths = [6, 18, 28, 80, 8, 7, 7, 7, 7, 9, 7, 7, 16]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[3].height = 35

# ───────── Sheet 3: 사업별 통계 ─────────
ws2 = wb.create_sheet('사업별')
ws2['A1'] = '사업별 요구사항 / Smell 검출'
ws2['A1'].font = TITLE_FONT
ws2.merge_cells('A1:F1')

with open(REPORT / 'requirements_per_project.csv', encoding='utf-8-sig') as f:
    pp = list(csv.reader(f))
HEADERS3 = ['No.', '기관', '사업명', '요구사항', 'Smell', '%']
for c, h in enumerate(HEADERS3, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = H_FONT; cell.fill = H_FILL
    cell.alignment = Alignment(horizontal='center')

for ri, row in enumerate(pp[1:], 4):
    pid, org, project, reqs, smelly, pct, *_ = row
    no = parse_pid(pid)
    cells = [no, org, project, int(reqs), int(smelly), float(pct)]
    for c, v in enumerate(cells, 1):
        cell = ws2.cell(row=ri, column=c, value=v)
        cell.font = CELL_FONT
        if c == 6: cell.number_format = '0.0"%"'
ws2.freeze_panes = 'A4'
for i, w in enumerate([6, 22, 48, 11, 11, 8], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ───────── Sheet 4: 필터 규칙 ─────────
ws3 = wb.create_sheet('필터 규칙')
ws3['A1'] = '요구사항 판별 규칙'
ws3['A1'].font = TITLE_FONT
ws3.merge_cells('A1:C1')
ws3['A3'] = '단계'; ws3['B3'] = '판정'; ws3['C3'] = '설명'
for col in 'ABC':
    c = ws3[f'{col}3']; c.font = H_FONT; c.fill = H_FILL
    c.alignment = Alignment(horizontal='center')

rules = [
    ('필수 종결어', 'PASS', '"~여야 한다", "~여야 함", "할 수 있어야 한다", "되도록 한다", "반드시 ~한다", "필수이다" 중 하나 매칭'),
    ('길이', 'CUT', '20자 미만 또는 350자 초과 컷'),
    ('사이트 푸터', 'CUT', 'COPYRIGHT, 패밀리사이트, 사이트맵, 페이지로 이동, 용어사전, 약관 등'),
    ('조달·입찰 안내', 'CUT', '입찰자, 공동수급체, 청렴계약, 조달청, 나라장터, 국가계약법, 법인등기부등본 등 조달 표준 문구'),
    ('메타데이터 시작', 'CUT', '사업명/사업기간/사업예산/문의처/연락처/발주기관 등으로 시작'),
    ('글머리표 시작', 'CUT', '▶ ○ ● ◇ □ ■, 로마숫자 Ⅰ~Ⅹ, "1." "1)" 등 단편 시작'),
    ('평가기준', 'CUT', '"~을 평가한다", "~을 점검한다" 등 RFP 채점 룰'),
    ('약관·제도 설명', 'CUT', '임차인/임대인/보증금/채권/법원 등 + 시스템 명사 없으면 컷'),
    ('사이트 공통 텍스트', 'CUT', '5개 이상 사업에 동일하게 등장 = 사이트 공통'),
    ('사업 내 중복', 'CUT', '같은 사업 안의 동일 문장 1회만 유지'),
]
for i, (a, b, c) in enumerate(rules, 4):
    ws3[f'A{i}'] = a; ws3[f'B{i}'] = b; ws3[f'C{i}'] = c
    for col in 'ABC':
        ws3[f'{col}{i}'].font = CELL_FONT
        ws3[f'{col}{i}'].alignment = Alignment(wrap_text=True, vertical='top')
    if b == 'CUT':
        ws3[f'B{i}'].fill = PatternFill('solid', fgColor='FFCCCC')
    else:
        ws3[f'B{i}'].fill = PatternFill('solid', fgColor='CCFFCC')

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 80

wb.save(OUT)
print(f'Excel: {OUT}')
print(f'크기: {OUT.stat().st_size / 1024:.0f} KB')
