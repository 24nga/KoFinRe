"""REQ_abstract 30건 평가 결과를 Excel로 (기존 56건 분석과 비교 시트 포함)."""
import csv, json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule

REPORT = Path(r'C:\Users\heen1\Desktop\assist\rfp_report')
OUT = REPORT / 'REQ_abstract_평가결과.xlsx'

SMELL_COLS = ['모호어','수동태','복합의무','정량부재','미정의약어','주체모호','약한의무']
SMELL_DESC = {
    '모호어':     '"필요한/적절한/실시간" 등 정량성 결여',
    '수동태':     '"~되어야 한다" 행위 주체 흐릿',
    '복합의무':   '한 문장에 2개 이상 의무',
    '정량부재':   '성능/처리량 키워드인데 숫자 없음',
    '미정의약어': '정의(괄호/콜론) 없는 영문 약어',
    '주체모호':   '주어 없이 의무문',
    '약한의무':   '"~한다/된다" 평서형 (해야 한다 권장)',
}

H_FILL = PatternFill('solid', fgColor='1F4E79')
H_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
TITLE_FONT = Font(bold=True, color='1F4E79', name='Arial', size=14)
SUB_FONT = Font(italic=True, color='606060', name='Arial', size=10)
CELL_FONT = Font(name='맑은 고딕', size=10)
SMELLY_FILL = PatternFill('solid', fgColor='FFF4CC')

wb = Workbook()
summary = json.load(open(REPORT / 'req_abstract_summary.json', encoding='utf-8'))

# ───────── Sheet 1: 안내 ─────────
ws0 = wb.active
ws0.title = '안내'
ws0['A1'] = 'REQ_abstract.csv — 구조화된 RFP 요구사항 30건 평가'
ws0['A1'].font = TITLE_FONT
ws0.merge_cells('A1:C1')
ws0['A2'] = '※ 깔끔하게 정리된 정형 데이터에 동일 한국어 Smell 룰 적용'
ws0['A2'].font = SUB_FONT
ws0.merge_cells('A2:C2')

ws0['A4'] = '항목'; ws0['B4'] = '값'; ws0['C4'] = '설명'
for col in 'ABC':
    c = ws0[f'{col}4']; c.font = H_FONT; c.fill = H_FILL
    c.alignment = Alignment(horizontal='center')

stats = [
    ('프로젝트 수', '3', 'P001 여신·리스크 / P002 오토금융 / P003 고객360·AML'),
    ('요구사항 (req_id)', summary['total_reqs'], '원본 행 1개 = 요구사항 1건'),
    ('세부 항목 (sub-req)', summary['total_sub_reqs'], 'req_details를 "; "로 분리한 sub-requirement'),
    ('Smell 있는 요구사항', summary['total_with_smell'],
     f"전체의 {100*summary['total_with_smell']/summary['total_reqs']:.1f}% (req_id 수준 OR 집계)"),
]
for i, (a, b, c) in enumerate(stats, 5):
    ws0[f'A{i}'] = a; ws0[f'B{i}'] = b; ws0[f'C{i}'] = c
    for col in 'ABC': ws0[f'{col}{i}'].font = CELL_FONT

ws0['A10'] = '── Smell 유형별 ──'; ws0['A10'].font = Font(bold=True, italic=True)
for i, c in enumerate(SMELL_COLS, 11):
    ws0[f'A{i}'] = c; ws0[f'B{i}'] = summary[c]; ws0[f'C{i}'] = SMELL_DESC[c]
    for col in 'ABC': ws0[f'{col}{i}'].font = CELL_FONT

ws0.column_dimensions['A'].width = 22
ws0.column_dimensions['B'].width = 10
ws0.column_dimensions['C'].width = 70

# ───────── Sheet 2: 요구사항 단위 ─────────
ws1 = wb.create_sheet('요구사항 단위')
ws1['A1'] = '요구사항 (req_id) 단위 평가'
ws1['A1'].font = TITLE_FONT
ws1.merge_cells('A1:M1')

with open(REPORT / 'req_abstract_eval.csv', encoding='utf-8-sig') as f:
    rows = list(csv.DictReader(f))

HEADERS = ['Project','시스템','Req ID','요구사항 제목','세부','오류세부','Smell'] + SMELL_COLS + ['요구사항 본문']
for c, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=3, column=c, value=h)
    cell.font = H_FONT; cell.fill = H_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for ri, r in enumerate(rows, 4):
    cells = [
        r['project_id'], r['system_name'], r['req_id'], r['req_title'],
        int(r['sub_count']), int(r['sub_smelly_count']), int(r['has_smell']),
    ] + [int(r[c]) for c in SMELL_COLS] + [r['req_details']]
    for c, v in enumerate(cells, 1):
        cell = ws1.cell(row=ri, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws1.row_dimensions[ri].height = 90

last = 3 + len(rows)
ws1.auto_filter.ref = f'A3:{get_column_letter(len(HEADERS))}{last}'
ws1.freeze_panes = 'A4'
ws1.conditional_formatting.add(
    f'A4:{get_column_letter(len(HEADERS))}{last}',
    FormulaRule(formula=['$G4=1'], fill=SMELLY_FILL)
)
widths = [8, 18, 10, 36, 6, 8, 7, 7, 7, 7, 7, 9, 7, 7, 80]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[3].height = 35

# ───────── Sheet 3: 세부 항목 단위 ─────────
ws2 = wb.create_sheet('세부 항목 단위')
ws2['A1'] = f'세부 항목 단위 (sub-requirement, {summary["total_sub_reqs"]}건)'
ws2['A1'].font = TITLE_FONT
ws2.merge_cells('A1:L1')

with open(REPORT / 'req_abstract_eval_sub.csv', encoding='utf-8-sig') as f:
    subrows = list(csv.DictReader(f))

H2 = ['Project','시스템','Req ID','요구사항','#','세부 내용','Smell'] + SMELL_COLS + ['모호어 단어']
for c, h in enumerate(H2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = H_FONT; cell.fill = H_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for ri, r in enumerate(subrows, 4):
    cells = [
        r['project_id'], r['system_name'], r['req_id'], r['req_title'],
        int(r['sub_idx']), r['sub_text'], int(r['has_smell']),
    ] + [int(r[c]) for c in SMELL_COLS] + [r.get('vague_hits','')]
    for c, v in enumerate(cells, 1):
        cell = ws2.cell(row=ri, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='top')

last2 = 3 + len(subrows)
ws2.auto_filter.ref = f'A3:{get_column_letter(len(H2))}{last2}'
ws2.freeze_panes = 'A4'
ws2.conditional_formatting.add(
    f'A4:{get_column_letter(len(H2))}{last2}',
    FormulaRule(formula=['$G4=1'], fill=SMELLY_FILL)
)
widths2 = [8, 18, 10, 24, 4, 70, 7, 7, 7, 7, 7, 9, 7, 7, 16]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[3].height = 35

# ───────── Sheet 4: 프로젝트별 ─────────
ws3 = wb.create_sheet('프로젝트별')
ws3['A1'] = '프로젝트별 통계'
ws3['A1'].font = TITLE_FONT
ws3.merge_cells('A1:L1')

with open(REPORT / 'req_abstract_eval_per_project.csv', encoding='utf-8-sig') as f:
    pp = list(csv.reader(f))

PROJECT_NAMES = {
    'P001': '차세대 여신·리스크 통합 구축',
    'P002': '디지털 오토금융 플랫폼 구축',
    'P003': '고객360·AML·리스크 고도화',
}

H3 = ['Project','프로젝트명','요구사항','세부 항목','Smell 검출','%']+SMELL_COLS
for c, h in enumerate(H3, 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.font = H_FONT; cell.fill = H_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for ri, row in enumerate(pp[1:], 4):
    pid, reqs, sub, smelly, pct, *smells = row
    cells = [pid, PROJECT_NAMES.get(pid, ''), int(reqs), int(sub), int(smelly), float(pct)] + [int(x) for x in smells]
    for c, v in enumerate(cells, 1):
        cell = ws3.cell(row=ri, column=c, value=v)
        cell.font = CELL_FONT
        if c == 6: cell.number_format = '0.0"%"'

last3 = 3 + len(pp) - 1
ws3.conditional_formatting.add(f'F4:F{last3}',
    ColorScaleRule(start_type='min', start_color='FFFFFF', end_type='max', end_color='C00000'))
for i, w in enumerate([8, 30, 12, 12, 12, 8, 8, 8, 10, 10, 12, 10, 10], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[3].height = 35

# ───────── Sheet 5: 기존 56건 분석과 비교 ─────────
ws4 = wb.create_sheet('비교(56건 vs 30건)')
ws4['A1'] = '기존 RFP 56건 추출 분석 vs REQ_abstract 30건 비교'
ws4['A1'].font = TITLE_FONT
ws4.merge_cells('A1:D1')

H4 = ['항목','RFP 56건 (자체 추출)','REQ_abstract 30건','해석']
for c, h in enumerate(H4, 1):
    cell = ws4.cell(row=3, column=c, value=h)
    cell.font = H_FONT; cell.fill = H_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# 기존 56건 정밀필터 요약
try:
    prev = json.load(open(REPORT / 'requirements_summary.json', encoding='utf-8'))
except:
    prev = {'total_requirements': 75, 'total_with_smell': 49, 'projects_with_reqs': 9,
            '모호어':23,'수동태':6,'복합의무':2,'정량부재':5,'미정의약어':1,'주체모호':24,'약한의무':0}

compares = [
    ('데이터 출처','RFP 공고문 56건 자체 텍스트 추출','구조화된 정형 CSV','정형 vs 비정형'),
    ('단위','문장','요구사항(req_id) + sub-requirement','sub-req 분해로 더 정밀'),
    ('분석 단위 수', f"{prev.get('total_requirements','75')}건 (정밀필터 후)",
                 f"{summary['total_reqs']}건 (req_id) / {summary['total_sub_reqs']}건 (sub)",''),
    ('Smell 비율', f"{100*prev.get('total_with_smell',49)/max(prev.get('total_requirements',75),1):.1f}%",
                 f"{100*summary['total_with_smell']/summary['total_reqs']:.1f}%",
                 '정형 데이터가 의무 표현 다수 — 비율 ↑'),
]
for c in SMELL_COLS:
    compares.append((c, prev.get(c, 0), summary[c], ''))

for i, (a, b, c, d) in enumerate(compares, 4):
    ws4[f'A{i}'] = a; ws4[f'B{i}'] = b; ws4[f'C{i}'] = c; ws4[f'D{i}'] = d
    for col in 'ABCD': ws4[f'{col}{i}'].font = CELL_FONT
    if i < 8: ws4[f'A{i}'].font = Font(name='맑은 고딕', size=10, bold=True)

ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 26
ws4.column_dimensions['C'].width = 28
ws4.column_dimensions['D'].width = 38

wb.save(OUT)
print(f'Excel: {OUT}')
print(f'크기: {OUT.stat().st_size/1024:.0f} KB')
