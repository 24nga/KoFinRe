"""분석된 모든 추출 문장 3,210건을 보기 좋은 Excel로."""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

REPORT = Path(r'C:\Users\heen1\Desktop\assist\rfp_report')
SRC = REPORT / 'sentences_all.csv'
OUT = REPORT / 'RFP_요구사항_원문_3210건.xlsx'

SMELL_COLS = ['모호어','수동태','복합의무','정량부재','미정의약어','주체모호','약한의무']

H_FILL = PatternFill('solid', fgColor='1F4E79')
H_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
TITLE_FONT = Font(bold=True, color='1F4E79', name='Arial', size=14)
SUB_FONT = Font(italic=True, color='606060', name='Arial', size=10)
CELL_FONT = Font(name='맑은 고딕', size=10)
SMELLY_FILL = PatternFill('solid', fgColor='FFF4CC')  # 연한 노랑

wb = Workbook()

# ───────── Sheet 1: 안내 ─────────
ws0 = wb.active
ws0.title = '안내'
ws0['A1'] = 'RFP 56건 — 추출된 요구사항 문장 원문 (3,210건)'
ws0['A1'].font = TITLE_FONT
ws0.merge_cells('A1:B1')
ws0['A2'] = '※ 56건의 RFP/공고문에서 한국어 비율·길이 필터로 추출한 모든 문장'
ws0['A2'].font = SUB_FONT
ws0.merge_cells('A2:B2')

ws0['A4'] = '시트 구성'; ws0['B4'] = '내용'
ws0['A4'].font = H_FONT; ws0['A4'].fill = H_FILL
ws0['B4'].font = H_FONT; ws0['B4'].fill = H_FILL

guides = [
    ('전체 문장', '3,210건 전부. has_smell=1 행은 노란색 배경. 자동 필터 적용 (열 머리글 ▼)'),
    ('검출 문장만', '771건 (smell 1개 이상)'),
    ('사업별 통계', '사업별 분석문장/검출 수'),
]
for i, (a, b) in enumerate(guides, 5):
    ws0[f'A{i}'] = a; ws0[f'B{i}'] = b
    ws0[f'A{i}'].font = CELL_FONT; ws0[f'B{i}'].font = CELL_FONT

ws0['A9']  = '컬럼 설명'; ws0['A9'].font = H_FONT; ws0['A9'].fill = H_FILL
ws0['B9']  = '의미'; ws0['B9'].font = H_FONT; ws0['B9'].fill = H_FILL

cols = [
    ('No.',         '사업번호 (01~56)'),
    ('기관',        '발주기관'),
    ('사업명',      '사업명'),
    ('문장',        '추출된 원문 (자르지 않음)'),
    ('has_smell',   '하나 이상의 smell 검출 (1/0)'),
    ('모호어',      '필요한·적절한·신속히 등'),
    ('수동태',      '~되어야 한다 어미'),
    ('복합의무',    '한 문장에 의무 표현 2개 이상'),
    ('정량부재',    '성능 키워드인데 숫자 없음'),
    ('미정의약어',  '정의 없는 영문 약어'),
    ('주체모호',    '주어 없이 의무문'),
    ('약한의무',    '~한다/된다 평서형 (해야 한다 권장)'),
    ('모호어 단어', '실제 매칭된 단어 목록 (|로 구분)'),
]
for i, (a, b) in enumerate(cols, 10):
    ws0[f'A{i}'] = a; ws0[f'A{i}'].font = CELL_FONT
    ws0[f'B{i}'] = b; ws0[f'B{i}'].font = CELL_FONT

ws0.column_dimensions['A'].width = 22
ws0.column_dimensions['B'].width = 80

# ───────── Sheet 2: 전체 문장 ─────────
ws1 = wb.create_sheet('전체 문장')
ws1['A1'] = '추출된 모든 분석 문장 (3,210건)'
ws1['A1'].font = TITLE_FONT
ws1.merge_cells('A1:M1')

HEADERS = ['No.', '기관', '사업명', '문장', 'has_smell'] + SMELL_COLS + ['모호어 단어']
for c, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=3, column=c, value=h)
    cell.font = H_FONT; cell.fill = H_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

def parse_pid(pid):
    return pid.split('_', 1)[0] if '_' in pid else pid

# 데이터 적재
rows = list(csv.DictReader(open(SRC, encoding='utf-8-sig')))
print(f'문장 로드: {len(rows)}건')

for ri, r in enumerate(rows, 4):
    cells = [
        parse_pid(r['project_id']),
        r['org'],
        r['project'],
        r['sentence'],
        int(r['has_smell']),
    ] + [int(r[c]) for c in SMELL_COLS] + [r.get('vague_hits', '')]
    for c, v in enumerate(cells, 1):
        cell = ws1.cell(row=ri, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# 자동 필터 + 첫 행 고정 + 노란 배경 조건부서식
last_row = 3 + len(rows)
ws1.auto_filter.ref = f'A3:M{last_row}'
ws1.freeze_panes = 'A4'

# has_smell=1 행 전체에 노란 배경
ws1.conditional_formatting.add(
    f'A4:M{last_row}',
    FormulaRule(formula=[f'$E4=1'], fill=SMELLY_FILL)
)

widths = [6, 18, 28, 75, 9, 8, 8, 8, 8, 10, 8, 8, 18]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[3].height = 35

# ───────── Sheet 3: 검출 문장만 ─────────
ws2 = wb.create_sheet('검출 문장만')
ws2['A1'] = 'Smell 검출 문장만 추린 것 (771건)'
ws2['A1'].font = TITLE_FONT
ws2.merge_cells('A1:M1')

for c, h in enumerate(HEADERS, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = H_FONT; cell.fill = H_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

smelly = [r for r in rows if r['has_smell'] == '1']
for ri, r in enumerate(smelly, 4):
    cells = [
        parse_pid(r['project_id']),
        r['org'], r['project'], r['sentence'],
        int(r['has_smell']),
    ] + [int(r[c]) for c in SMELL_COLS] + [r.get('vague_hits', '')]
    for c, v in enumerate(cells, 1):
        cell = ws2.cell(row=ri, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='top')

last2 = 3 + len(smelly)
ws2.auto_filter.ref = f'A3:M{last2}'
ws2.freeze_panes = 'A4'
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[3].height = 35

# ───────── Sheet 4: 사업별 통계 ─────────
ws3 = wb.create_sheet('사업별 통계')
ws3['A1'] = '사업별 분석 문장 / Smell 검출'
ws3['A1'].font = TITLE_FONT
ws3.merge_cells('A1:F1')

HEADERS3 = ['No.', '기관', '사업명', '분석 문장', 'Smell 검출', '%']
for c, h in enumerate(HEADERS3, 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.font = H_FONT; cell.fill = H_FILL
    cell.alignment = Alignment(horizontal='center')

with open(REPORT / 'per_project.csv', encoding='utf-8-sig') as f:
    pp = list(csv.reader(f))
for ri, row in enumerate(pp[1:], 4):
    pid, org, project, sentences, with_smell, pct, *_ = row
    no = parse_pid(pid)
    cells = [no, org, project, int(sentences), int(with_smell), float(pct)]
    for c, v in enumerate(cells, 1):
        cell = ws3.cell(row=ri, column=c, value=v)
        cell.font = CELL_FONT
        if c == 6: cell.number_format = '0.0"%"'
ws3.freeze_panes = 'A4'
for i, w in enumerate([6, 22, 48, 11, 11, 8], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)
print(f'Excel: {OUT}')
print(f'크기: {OUT.stat().st_size / 1024:.0f} KB')
