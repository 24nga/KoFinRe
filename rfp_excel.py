"""RFP 56건 한국어 smell 분석 → Excel 리포트."""
import csv, json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

REPORT = Path(r'C:\Users\heen1\Desktop\assist\rfp_report')
OUT = REPORT / 'RFP_56건_한국어_Smell분석.xlsx'

SMELL_COLS = ['모호어','수동태','복합의무','정량부재','미정의약어','주체모호','약한의무']
SMELL_DESC = {
    '모호어':     '"필요한·적절한·신속히·효율적" 등 정량성 결여 표현',
    '수동태':     '"~되어야 한다" — 행위 주체가 흐릿함',
    '복합의무':   '한 문장에 2개 이상의 의무 표현 → 분리 필요',
    '정량부재':   '성능·규모 요구에 숫자/단위 없음',
    '미정의약어': '정의(괄호·콜론) 없이 등장하는 영문 약어',
    '주체모호':   '주어(~가/이/은/는) 없이 의무문 시작',
    '약한의무':   '"한다/된다" 평서형 — "해야 한다"가 권장',
}

wb = Workbook()
H_FILL = PatternFill('solid', fgColor='1F4E79')
H_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
TITLE_FONT = Font(bold=True, color='1F4E79', name='Arial', size=14)
SUB_FONT = Font(italic=True, color='606060', name='Arial', size=10)
CELL_FONT = Font(name='맑은 고딕', size=10)

# ───────── Sheet 1: 전체 요약 ─────────
ws1 = wb.active
ws1.title = '전체 요약'
ws1['A1'] = 'RFP 56건 한국어 Smell 분석 — 전체 요약'
ws1['A1'].font = TITLE_FONT
ws1.merge_cells('A1:C1')
ws1['A2'] = '※ Paska(영문)와 동일한 분석 구조를 한국어 RFP에 맞춰 자체 구현'
ws1['A2'].font = SUB_FONT
ws1.merge_cells('A2:C2')

# Headers
for col, h in zip('ABC', ['항목', '값', '설명']):
    cell = ws1[f'{col}4']
    cell.value = h; cell.font = H_FONT; cell.fill = H_FILL
    cell.alignment = Alignment(horizontal='center')

# Load summary
summary = json.load(open(REPORT / 'summary.json', encoding='utf-8'))

r = 5
ws1[f'A{r}'] = '분석 사업 수'; ws1[f'B{r}'] = 56; ws1[f'C{r}'] = '원본 56건 전부 시도'
r += 1
ws1[f'A{r}'] = '본문 추출 양호한 사업'; ws1[f'B{r}'] = '14건'; ws1[f'C{r}'] = '나머지는 사이트 차단/인증 필요로 빈 응답'
r += 1
ws1[f'A{r}'] = '분석 가능 문장'; ws1[f'B{r}'] = summary['total_sentences']; ws1[f'C{r}'] = '문장 분리 + 한국어 비율 ≥25% 필터 통과'
r += 1
ws1[f'A{r}'] = 'Smell 검출 문장'; ws1[f'B{r}'] = summary['total_with_smell']
pct = 100 * summary['total_with_smell'] / max(summary['total_sentences'], 1)
ws1[f'C{r}'] = f'전체의 {pct:.1f}%'
r += 2
ws1[f'A{r}'] = '─── Smell 유형별 ───'
ws1[f'A{r}'].font = Font(bold=True, italic=True, name='Arial')
r += 1
for k in SMELL_COLS:
    ws1[f'A{r}'] = k
    ws1[f'B{r}'] = summary[k]
    ws1[f'C{r}'] = SMELL_DESC[k]
    r += 1

# 폭
for col, w in zip('ABC', [24, 12, 70]):
    ws1.column_dimensions[col].width = w
for row in range(5, r):
    for col in 'ABC':
        ws1[f'{col}{row}'].font = CELL_FONT

# ───────── Sheet 2: 사업별 통계 ─────────
ws2 = wb.create_sheet('사업별 통계')
ws2['A1'] = '사업별 Smell 검출 현황 (사업 56건)'
ws2['A1'].font = TITLE_FONT
ws2.merge_cells('A1:L1')

with open(REPORT / 'per_project.csv', encoding='utf-8-sig') as f:
    rows = list(csv.reader(f))

# Reorder columns to be readable
# CSV: project_id, org, project, sentences, with_smell, % smelly, 모호어 ... 약한의무
HEADERS = ['No.', '기관', '사업명', '분석문장', 'Smell검출', '%'] + SMELL_COLS
for c, h in enumerate(HEADERS, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = H_FONT; cell.fill = H_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# project_id 형식: NN_orgIdx_orgName_NN_year_projname  → 사업번호 추출
def parse_pid(pid):
    parts = pid.split('_', 3)
    return parts[0] if parts else pid

for ri, row in enumerate(rows[1:], 4):
    pid, org, project, sentences, with_smell, pct, *smells = row
    no = parse_pid(pid)
    cells = [no, org, project, int(sentences), int(with_smell), float(pct)] + [int(x) for x in smells]
    for c, v in enumerate(cells, 1):
        cell = ws2.cell(row=ri, column=c, value=v)
        cell.font = CELL_FONT
        if c == 6: cell.number_format = '0.0"%"'

# 색상 스케일 % 열
last_row = 3 + len(rows) - 1
ws2.conditional_formatting.add(f'F4:F{last_row}',
    ColorScaleRule(start_type='min', start_color='FFFFFF',
                   end_type='max', end_color='C00000'))

# 폭
widths = [6, 22, 45, 10, 10, 8, 10, 10, 10, 10, 12, 10, 10]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[3].height = 35
ws2.freeze_panes = 'D4'

# ───────── Sheet 3: 검출 문장 상세 ─────────
ws3 = wb.create_sheet('검출 문장 상세')
ws3['A1'] = f'Smell이 검출된 문장 상세 ({summary["total_with_smell"]}건)'
ws3['A1'].font = TITLE_FONT
ws3.merge_cells('A1:L1')

with open(REPORT / 'smell.csv', encoding='utf-8-sig') as f:
    rows = list(csv.reader(f))

# CSV columns: project_id, org, project, sentence, 모호어..약한의무, vague_hits
HEADERS3 = ['No.', '기관', '사업명', '문장'] + SMELL_COLS + ['모호어 단어']
for c, h in enumerate(HEADERS3, 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.font = H_FONT; cell.fill = H_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for ri, row in enumerate(rows[1:], 4):
    pid, org, project, sentence, *rest = row
    smells = rest[:-1]
    vague = rest[-1]
    no = parse_pid(pid)
    cells = [no, org, project, sentence] + [int(x) for x in smells] + [vague]
    for c, v in enumerate(cells, 1):
        cell = ws3.cell(row=ri, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='top')

widths = [6, 18, 30, 70, 8, 8, 8, 8, 10, 8, 8, 20]
for i, w in enumerate(widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[3].height = 35
ws3.freeze_panes = 'A4'

# ───────── Sheet 4: 키워드 Top ─────────
ws4 = wb.create_sheet('자주 등장 단어')
ws4['A1'] = '자주 등장하는 모호어 / 미정의 약어 Top 20'
ws4['A1'].font = TITLE_FONT
ws4.merge_cells('A1:E1')

ws4['A3'] = '모호어'; ws4['B3'] = '빈도'
ws4['D3'] = '미정의 약어'; ws4['E3'] = '빈도'
for col in 'ABDE':
    ws4[f'{col}3'].font = H_FONT; ws4[f'{col}3'].fill = H_FILL
    ws4[f'{col}3'].alignment = Alignment(horizontal='center')

for i, (term, n) in enumerate(summary['vague_top'][:20], 4):
    ws4[f'A{i}'] = term; ws4[f'A{i}'].font = CELL_FONT
    ws4[f'B{i}'] = n;    ws4[f'B{i}'].font = CELL_FONT
for i, (term, n) in enumerate(summary['undefined_acronym_top'][:20], 4):
    ws4[f'D{i}'] = term; ws4[f'D{i}'].font = CELL_FONT
    ws4[f'E{i}'] = n;    ws4[f'E{i}'].font = CELL_FONT

for col, w in zip('ABCDE', [18, 10, 4, 18, 10]):
    ws4.column_dimensions[col].width = w

# ───────── Sheet 5: 분석 방식 설명 ─────────
ws5 = wb.create_sheet('분석 방식')
ws5['A1'] = '한국어 RFP Smell 분석 — 점검 항목 정의'
ws5['A1'].font = TITLE_FONT
ws5.merge_cells('A1:D1')

ws5['A3'] = '점검 항목'; ws5['B3'] = '대응 Paska smell'; ws5['C3'] = '판정 규칙'; ws5['D3'] = '예시'
for col in 'ABCD':
    ws5[f'{col}3'].font = H_FONT; ws5[f'{col}3'].fill = H_FILL
    ws5[f'{col}3'].alignment = Alignment(horizontal='center')

RULES = [
    ('모호어','Coordination ambiguity / Not precise verb',
     '필요한·적절한·신속히·효율적 등 30+종 단어 매칭',
     '"시스템은 적절한 응답시간을 보장한다"'),
    ('수동태','Passive voice',
     '"~되어야 한다", "~지원되어야" 등 어미 매칭',
     '"보고서는 자동으로 생성되어야 한다"'),
    ('복합의무','Non-atomic requirement',
     '한 문장에 "해야 한다/할 수 있어야" 2회 이상',
     '"A를 처리해야 하고 B를 전송해야 한다"'),
    ('정량부재','Incomplete requirement',
     '성능·규모 키워드(처리량·TPS·용량 등) 있는데 숫자·단위 없음',
     '"빠른 응답속도를 보장한다"'),
    ('미정의약어','— (Paska 미커버, RFP 특화)',
     '문서 전체에서 같은 약어의 정의(괄호/콜론) 부재',
     '"AML/CFT 처리를 위해…" (AML 첫 등장)'),
    ('주체모호','Incomplete system response',
     '주어 없이 의무문 시작 (가/이/은/는 부재)',
     '"매월 정기적으로 보고한다"'),
    ('약한의무','Not requirement (반대 신호)',
     '"~한다/된다" 평서형 + 강한 의무 표현 없음',
     '"이력 정보를 저장한다" (해야 한다 권장)'),
]
for i, (name, mapping, rule, ex) in enumerate(RULES, 4):
    ws5[f'A{i}'] = name; ws5[f'B{i}'] = mapping; ws5[f'C{i}'] = rule; ws5[f'D{i}'] = ex
    for col in 'ABCD':
        ws5[f'{col}{i}'].font = CELL_FONT
        ws5[f'{col}{i}'].alignment = Alignment(wrap_text=True, vertical='top')

ws5.column_dimensions['A'].width = 14
ws5.column_dimensions['B'].width = 38
ws5.column_dimensions['C'].width = 45
ws5.column_dimensions['D'].width = 42
ws5.row_dimensions[3].height = 30

wb.save(OUT)
print(f'Excel 리포트: {OUT}')
