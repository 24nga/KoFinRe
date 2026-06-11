"""Excel 다중 시트 리포트 작성 (v2.1 핵심 정리)."""
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


H_FILL = PatternFill('solid', fgColor='1F4E79')
H_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
TITLE_FONT = Font(bold=True, color='1F4E79', name='Arial', size=14)
CELL_FONT = Font(name='맑은 고딕', size=10)
SMELLY_FILL = PatternFill('solid', fgColor='FFF4CC')

SMELL_CODES_KO = {
    "S1":"복합의무","S2":"불완전","S3":"모호어","S4":"약한의무","S5":"주체누락",
    "S6":"정량부재","S7":"미정의약어","S8":"범위모호","S9":"수동표현","S10":"검증불가",
}


def _write_sheet(ws, title: str, headers: List[str], rows: List[List[Any]],
                 widths: List[int] = None, highlight_smell_col: int = None):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1,
                  end_row=1, end_column=len(headers))
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = H_FONT; cell.fill = H_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for ri, row in enumerate(rows, 4):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = CELL_FONT
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 35
    if rows:
        last = 3 + len(rows)
        ws.auto_filter.ref = f'A3:{get_column_letter(len(headers))}{last}'
        ws.freeze_panes = 'A4'
        if highlight_smell_col is not None:
            col_letter = get_column_letter(highlight_smell_col)
            ws.conditional_formatting.add(
                f'A4:{get_column_letter(len(headers))}{last}',
                FormulaRule(formula=[f'${col_letter}4=1'], fill=SMELLY_FILL)
            )


def write_xlsx_report(out_path: Path, *,
                       summary: Dict[str, Any] = None,
                       requirements: List[Dict[str, Any]] = None,
                       per_project: List[Dict[str, Any]] = None,
                       per_detector_metrics: Dict[str, Dict[str, float]] = None):
    """공통 Excel 리포트 — 4 시트 (안내 / 요구사항 / 프로젝트별 / 평가).

    Args:
        summary: 전체 통계
        requirements: 요구사항별 결과 [{project_id, req_id, sentence, has_smell, S1..S10}]
        per_project: 프로젝트별 통계
        per_detector_metrics: 평가 결과 (gold label 있을 때)
    """
    wb = Workbook()

    # Sheet 1 — 안내·요약
    ws0 = wb.active
    ws0.title = '안내'
    ws0['A1'] = 'KoFinRe-QA Analysis Report'
    ws0['A1'].font = TITLE_FONT
    ws0.merge_cells('A1:B1')
    r = 3
    if summary:
        for k, v in summary.items():
            ws0.cell(row=r, column=1, value=k).font = CELL_FONT
            ws0.cell(row=r, column=2, value=str(v)).font = CELL_FONT
            r += 1
    ws0.column_dimensions['A'].width = 28
    ws0.column_dimensions['B'].width = 60

    # Sheet 2 — 요구사항별
    if requirements:
        ws1 = wb.create_sheet('요구사항')
        headers = ['Project', 'Req ID', '문장', 'has_smell'] + list(SMELL_CODES_KO.keys())
        rows = []
        for r in requirements:
            row = [
                r.get('project_id', ''),
                r.get('req_id', ''),
                r.get('sentence', '')[:300],
                int(r.get('has_smell', 0)),
            ] + [int(r.get(c, 0)) for c in SMELL_CODES_KO]
            rows.append(row)
        widths = [10, 12, 70, 9] + [7] * 10
        _write_sheet(ws1, '요구사항별 평가', headers, rows, widths,
                    highlight_smell_col=4)

    # Sheet 3 — 프로젝트별
    if per_project:
        ws2 = wb.create_sheet('프로젝트별')
        headers = list(per_project[0].keys())
        rows = [[r.get(h, '') for h in headers] for r in per_project]
        _write_sheet(ws2, '프로젝트별 통계', headers, rows, [12] * len(headers))

    # Sheet 4 — Detection 성능
    if per_detector_metrics:
        ws3 = wb.create_sheet('평가 성능')
        headers = ['Smell', 'Precision', 'Recall', 'F1', 'FPR', 'FNR']
        rows = []
        for code, m in per_detector_metrics.items():
            rows.append([code,
                        round(m.get('precision', 0), 3),
                        round(m.get('recall', 0), 3),
                        round(m.get('f1', 0), 3),
                        round(m.get('fpr', 0), 3),
                        round(m.get('fnr', 0), 3)])
        _write_sheet(ws3, 'Detection Performance', headers, rows,
                    [10, 12, 12, 12, 12, 12])

    wb.save(out_path)
