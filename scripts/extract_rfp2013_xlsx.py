"""RFP 2013 데이터셋의 조달청 XLSX → sentence_candidates 형식 변환.

XLSX 시트별 (기능/성능/인터페이스/데이터/테스트/보안/품질/제약/PMR/PSR) 요구사항을
'요구사항명: 요구사항 내용' 형식으로 합쳐 CSV로 출력.
"""
import csv
import sys
from pathlib import Path

import openpyxl


XLSX = Path(r'data/rfp_2013/SW개발사업/[조달청]홈페이지_재구축사업_요구사항세부내용.xlsx')


def find_header_row(ws, max_scan=10):
    for r in range(1, min(max_scan, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and ('요구사항ID' in str(v) or '요구항목ID' in str(v)):
                return r
    return None


def col_index(ws, header_row, *candidates):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if not v:
            continue
        vs = str(v).strip()
        for cand in candidates:
            if cand in vs:
                return c
    return None


def main(out_csv: Path):
    if not XLSX.exists():
        print(f'NOT FOUND: {XLSX}')
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    rows = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        hr = find_header_row(ws)
        if not hr:
            print(f'[skip] {sn}: 헤더 없음')
            continue

        id_col = col_index(ws, hr, '요구사항ID', '요구항목ID')
        name_col = col_index(ws, hr, '요구항목명칭', '요구사항명')
        desc_col = col_index(ws, hr, '요구사항 내용', '요구내용', '내용')

        if not id_col:
            print(f'[skip] {sn}: ID 컬럼 없음')
            continue

        count = 0
        for r in range(hr + 1, ws.max_row + 1):
            rid = ws.cell(r, id_col).value if id_col else None
            name = ws.cell(r, name_col).value if name_col else ''
            desc = ws.cell(r, desc_col).value if desc_col else ''
            if not rid:
                continue
            name = str(name or '').strip()
            desc = str(desc or '').strip()
            text = name + (': ' + desc if desc and desc != '0' else '')
            text = text.strip(' :').strip()
            if len(text) < 5:
                continue
            rows.append({
                'doc_id': f'[조달청]홈페이지_{sn}',
                'sentence': f'[{rid}] {text}',
            })
            count += 1
        print(f'[{sn}] {count} req')

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with open(out_csv, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['doc_id', 'sentence'])
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print(f'\n총 {len(rows)} req → {out_csv}')


if __name__ == '__main__':
    out = Path(sys.argv[1] if len(sys.argv) > 1
               else 'results/rfp_2013/stage1/xlsx_requirements.csv')
    main(out)
