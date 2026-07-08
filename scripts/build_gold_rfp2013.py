"""D-MAIN(RFP 2013) 골드라벨링 시트 생성 — KCI_DRAFT.md §4.2 프로토콜 구현.

- 입력: experiments/rfp_2013_sample/stage3/all_detection.csv (공개·재현 가능)
- 표본: 층화 무작위 200건 (문서 형태 비례 HWP 167 / XLSX 33, seed=42)
- 재라벨: 200건 중 50건 (seed=43, 순서 셔플) — 최소 2주 후 intra-rater kappa용
- 블라인드: 도구 판정 플래그는 시트에 포함하지 않음 (앵커링 방지)
- 출력: results/rfp_2013/gold/gold_labeling_sheet.xlsx (3 시트: 가이드/본라벨200/재라벨50)

사용: python scripts/build_gold_rfp2013.py
"""
import csv
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT = Path('experiments/rfp_2013_sample/stage3/all_detection.csv')
OUT = Path('results/rfp_2013/gold/gold_labeling_sheet.xlsx')

SEED_MAIN, SEED_RELABEL = 42, 43
N_TOTAL, N_XLSX, N_RELABEL = 200, 33, 50

SMELLS = [
    ('S1', '복합의무', '한 문장에 의무 표현(해야 한다 등) 2회 이상 — 둘 이상 기능 결합', '로그인과 권한관리를 지원해야 하며 이력도 저장해야 한다'),
    ('S2', '불완전', '의무 표현은 있으나 행위 대상·시스템 응답이 빠짐', '~할 경우 처리하여야 한다 (무엇을?)'),
    ('S3', '모호어', '"적절히·필요한·실시간·효율적" 등 해석 불명 어휘', '적절한 성능을 확보해야 한다'),
    ('S4', '약한의무', '평서형 종결("한다/된다/함/됨")로 의무성 불명', '이력 정보를 저장한다'),
    ('S5', '주체누락', '수행/책임 주체(시스템·사용자·수행사 등) 미명시', '(주어 없이) 데이터를 암호화하여야 한다'),
    ('S6', '정량부재', '성능·용량 키워드가 있으나 수치 기준 없음', '응답시간이 빨라야 한다'),
    ('S7', '미정의약어', '정의 없이 사용된 영문 약어 (일반 약어 제외)', 'COSEQ 연계를 지원해야 한다'),
    ('S8', '범위모호', '"및/또는, 등, 포함, 관련"으로 범위 불명', '인증, 권한관리 등을 지원해야 한다'),
    ('S9', '수동표현', '"~되어야 한다" 계열로 행위 주체 흐림', '데이터가 관리되어야 한다'),
    ('S10', '검증불가', '정성 표현 + 판정 기준 부재', '안정적으로 운영되어야 한다'),
    ('S11', '구현편향', '특정 기술·제품·언어를 불필요하게 강제 (WHAT 아닌 HOW)', 'Java Spring으로 구현해야 한다'),
    ('S12', '부정문', '부정형 의무 ("~지 않아야/안 된다") — 검증 곤란', '오류가 발생하지 않아야 한다'),
    ('S13', '추측표현', '"가능하다면·추후 결정·필요시" 등 미확정', '필요시 알림을 발송할 수 있다'),
    ('S14', '수혜자불명', '의무문이나 수혜 이해관계자(사용자·운영자 등) 부재', '통계를 생성하여야 한다 (누구를 위해?)'),
    ('S15', '지시어모호', '"해당·상기·그것·본 건" 등 선행어 불명 지시', '해당 기능은 상기 조건을 따른다'),
    ('S16', '필요성불명확', '선호·임의·강제 등 사업 정당성 없는 요구', '개발팀이 선호하는 폰트를 강제로 사용'),
    ('S17', '실현불가', '"100% 가용/0초/완벽한/99.9999%" 등 실현 불가 절대치', '100% 가용성을 보장해야 한다'),
    ('S18', '추적ID부재', '강한 의무 표현 + 요구사항 ID 부재 + 출처·근거 인용 부재', '(ID 없이) 본 시스템은 인증을 제공하여야 한다'),
    ('S19', '제약분류불명', '제약·제한 시그널은 있으나 기술/사업/법규/운영/보안 어느 유형인지 불명', '본 사업은 일부 제약을 받아야 한다'),
]

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(color='FFFFFF', bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')


def load_rows():
    with open(INPUT, encoding='utf-8-sig') as f:
        rows = list(csv.DictReader(f))
    for i, r in enumerate(rows):
        r['_idx'] = i
    return rows


def stratified_sample(rows):
    xlsx = [r for r in rows if r['source'] == 'XLSX 요구사항']
    hwp = [r for r in rows if r['source'] != 'XLSX 요구사항']
    rnd = random.Random(SEED_MAIN)
    picked = rnd.sample(xlsx, N_XLSX) + rnd.sample(hwp, N_TOTAL - N_XLSX)
    rnd.shuffle(picked)
    return picked


def write_guide(ws):
    ws.append(['골드라벨링 가이드 — KCI_DRAFT.md §4.2 프로토콜'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    for line in [
        '1. 블라인드 원칙: 도구 판정을 보지 말고 문장만 읽고 판단한다.',
        '2. 각 스멜 컬럼에 해당하면 1, 아니면 0 (빈칸은 0으로 간주).',
        '3. 판단 곤란 시 메모 컬럼에 사유를 남기고 보수적으로 0.',
        '4. 본라벨 200건 완료 후 최소 2주 뒤 "재라벨50" 시트를 독립적으로 라벨링한다 (본라벨 결과 열람 금지).',
        '5. 세션당 50건 이하 권장 (피로 편향 방지).',
    ]:
        ws.append([line])
    ws.append([])
    ws.append(['Code', '명칭', '판정 기준', '양성 예'])
    for c in range(1, 5):
        cell = ws.cell(ws.max_row, c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    for code, name, crit, ex in SMELLS:
        ws.append([code, name, crit, ex])
    for col, w in zip('ABCD', [8, 14, 62, 44]):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=10):
        for cell in row:
            cell.alignment = WRAP


def write_label_sheet(ws, sample, id_prefix):
    headers = ['sample_id', 'doc_id', 'source', 'sentence'] + [s[0] for s in SMELLS] + ['메모']
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    for n, r in enumerate(sample, 1):
        ws.append([f'{id_prefix}{n:03d}', r['doc_id'][:40], r['source'], r['sentence']]
                  + [''] * len(SMELLS) + [''])
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 90
    for i in range(len(SMELLS)):
        ws.column_dimensions[get_column_letter(5 + i)].width = 5
    ws.column_dimensions[get_column_letter(5 + len(SMELLS))].width = 30
    for row in ws.iter_rows(min_row=2):
        row[3].alignment = WRAP
    ws.freeze_panes = 'E2'


def main():
    rows = load_rows()
    sample = stratified_sample(rows)

    rnd2 = random.Random(SEED_RELABEL)
    relabel = rnd2.sample(sample, N_RELABEL)
    rnd2.shuffle(relabel)

    wb = Workbook()
    write_guide(wb.active)
    wb.active.title = '가이드'
    write_label_sheet(wb.create_sheet('본라벨200'), sample, 'G')
    write_label_sheet(wb.create_sheet('재라벨50'), relabel, 'R')

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)

    # 매핑 파일 (채점용 — 라벨링 중에는 열지 말 것)
    with open(OUT.parent / 'sample_index_map.csv', 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(['sample_id', 'all_detection_row_idx'])
        for n, r in enumerate(sample, 1):
            w.writerow([f'G{n:03d}', r['_idx']])
    n_x = sum(1 for r in sample if r['source'] == 'XLSX 요구사항')
    print(f'본라벨 {len(sample)}건 (XLSX {n_x} / HWP {len(sample)-n_x}) + 재라벨 {len(relabel)}건')
    print(f'→ {OUT}')


if __name__ == '__main__':
    main()
